# -*- coding: utf-8 -*-
"""
Created on Sun Jan 27 16:03:36 2019

@author: Ajit
"""

import openpyxl
import datetime
today=datetime.datetime.today().strftime('%d-%m-%Y')

wb=openpyxl.load_workbook('C:\\Users\\Ajit\\01_02_2019.xlsx',data_only=True)
sheets=wb.sheetnames
sheet1=wb[sheets[1]]#1 or 4
sheet2=wb[sheets[2]]#2 or 5
col_names=[]
for rows in sheet1.iter_rows(min_row=1,max_row=1,min_col=355,max_col=706):#355,706 or 2,352
    for cell in rows:
        col_names.append(cell.value)
j=1
for names in col_names:
    sheet2.cell(1,j).value=names
    j=j+1
sheet2.cell(2,1).value="P(2|1)"
sheet2.cell(3,1).value="P(3|2)"
sheet2.cell(4,1).value="P(4|3)"
sheet2.cell(5,1).value="P(5|4)"
sheet2.cell(6,1).value="P(6|5)"
sheet2.cell(7,1).value="P(7|6)"
sheet2.cell(8,1).value="P(2-|1)"
sheet2.cell(9,1).value="P(3-|2)"
sheet2.cell(10,1).value="P(4-|3)"
sheet2.cell(11,1).value="P(5-|4)"
sheet2.cell(12,1).value="P(6-|5)"
sheet2.cell(13,1).value="P(7-|6)"
sheet2.cell(14,1).value="P(2|1-)"
sheet2.cell(15,1).value="P(3|2-)"
sheet2.cell(16,1).value="P(4|3-)"
sheet2.cell(17,1).value="P(5|4-)"
sheet2.cell(18,1).value="P(6|5-)"
sheet2.cell(19,1).value="P(7|6-)"
sheet2.cell(20,1).value="P(++-)"
sheet2.cell(21,1).value="P(-+-)"
sheet2.cell(22,1).value="P(+-+)"
sheet2.cell(23,1).value="P(--+)"
sheet2.cell(24,1).value="avg2_1"
sheet2.cell(25,1).value="avg3_2"
sheet2.cell(26,1).value="avg4_3"
sheet2.cell(27,1).value="avg5_4"
sheet2.cell(28,1).value="avg6_5"
sheet2.cell(29,1).value="avg7_6"

for cols in range(355,706):#355,706
    returns=[]
    for rows in range(2,502):
        a=str(sheet1.cell(row=rows,column=cols).value)
        a=a.replace(',','')
        a=float(a)
        returns.append(a)
    returns.reverse()
    N_1pos=0
    
    for num in returns:
        if(num>0):
            N_1pos+=1
    
    
    avg2=0
    avg3=0
    avg4=0
    avg5=0
    avg6=0
    avg7=0
    
    N_2pos=0
    
    i=1#P(2|1)
    for num in returns:
        if(i<500):
            if(returns[i]>0 and num>0):
                N_2pos+=1
                avg2=avg2+returns[i]
        i=i+1
    
    i=2#P(3|2)
    N_3pos=0
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-1]>0 and returns[i]>0):
                N_3pos+=1
                avg3=avg3+returns[i]
        i=i+1
            
    N_4pos=0
    i=3
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]>0):
                N_4pos+=1
                avg4=avg4+returns[i]
        i=i+1
    N_5pos=0
    i=4
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]>0):
                N_5pos+=1
                avg5=avg5+returns[i]
        i=i+1
    N_6pos=0
    i=5
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-4]>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]>0):
                N_6pos+=1
                avg6+=returns[i]
        i=i+1
    N_7pos=0
    i=6
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-5]>0 and returns[i-4]>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]>0):
                N_7pos+=1
                avg7+=returns[i]
        i=i+1
    
#    print("N_1pos=%s %s %s %s %s %s %s"%(N_1pos,N_2pos,N_3pos,N_4pos,N_5pos,N_6pos,N_7pos))
#    ###############################################################################################################
    N_1neg=500-N_1pos
    
    N_2pos_1neg=0
    i=1
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i]>0):
                N_2pos_1neg+=1
        i=i+1
    
    N_3pos_2neg=0
    i=2
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-1]<=0 and returns[i]>0):
                N_3pos_2neg+=1
        i=i+1
    
    N_4pos_3neg=0
    i=3
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]>0):
                N_4pos_3neg+=1
        i=i+1
    
    N_5pos_4neg=0
    i=4
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-3]<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]>0):
                N_5pos_4neg+=1
        i=i+1
    N_6pos_5neg=0
    i=5
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-4]<=0 and returns[i-3]<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]>0):
                N_6pos_5neg+=1
        i=i+1
    N_7pos_6neg=0
    i=6
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-5]<=0 and returns[i-4]<=0 and returns[i-3]<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]>0):
                N_7pos_6neg+=1
        i=i+1
#    print("N_2pos_1neg=%s %s %s %s %s %s"%(N_2pos_1neg,N_3pos_2neg,N_4pos_3neg,N_5pos_4neg,N_6pos_5neg,N_7pos_6neg))
#    ########################################################################################################
    
    N_2neg_1pos=0
    i=1
    for num in returns:
        if(i<500):
            if(num>0 and returns[i]<=0):
                N_2neg_1pos+=1
        i=i+1
    
    N_3neg_2pos=0
    i=2
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-1]>0 and returns[i]<=0):
                N_3neg_2pos+=1
        i=i+1
    
    N_4neg_3pos=0
    i=3
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]<=0):
                N_4neg_3pos+=1
        i=i+1
    
    N_5neg_4pos=0
    i=4
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]<=0):
                N_5neg_4pos+=1
        i=i+1
    N_6neg_5pos=0
    i=5
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-4]>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]<=0):
                N_6neg_5pos+=1
        i=i+1
    N_7neg_6pos=0
    i=6
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-5]>0 and returns[i-4]>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]<=0):
                N_7neg_6pos+=1
        i=i+1
#    print("N_2neg_1pos=%s %s %s %s %s %s"%(N_2neg_1pos,N_3neg_2pos,N_4neg_3pos,N_5neg_4pos,N_6neg_5pos,N_7neg_6pos))
#   
    i=2#P(3|2)
    N_pos_pos_neg=0
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-1]>0 and returns[i]<=0):
                N_pos_pos_neg+=1
            i=i+1
    try:
        P_pos_pos_neg=N_pos_pos_neg/498
    except:
        P_pos_pos_neg=0

    i=2#P(3|2)
    N_neg_pos_neg=0
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-1]>0 and returns[i]<=0):
                N_neg_pos_neg+=1
        i=i+1
    try:
        P_neg_pos_neg=N_neg_pos_neg/498
    except:
        P_neg_pos_neg=0
    i=2#P(3|2)
    N_pos_neg_pos=0
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-1]>0 and returns[i]>0):
                N_pos_neg_pos+=1
            i=i+1
    try:
        P_pos_neg_pos=N_pos_neg_pos/498
    except:
        P_pos_neg_pos=0
    i=2#P(3|2)
    N_neg_neg_pos=0
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-1]<=0 and returns[i]>0):
                N_neg_neg_pos+=1
        i=i+1
    try:
        P_neg_neg_pos=N_neg_neg_pos/498
    except:
        P_neg_neg_pos=0
#################################################################################################################
    try:
        P_2pos_1pos=(N_2pos/499)/(N_1pos/500)
    except:
        P_2pos_1pos=0
    try:
        P_3pos_2pos=(N_3pos/498)/(N_2pos/499)
    except:
        P_3pos_2pos=0
    try:
        P_4pos_3pos=(N_4pos/497)/(N_3pos/498)
    except:
        P_4pos_3pos=0
    try:
        
        P_5pos_4pos=(N_5pos/496)/(N_4pos/497)
    except:
        P_5pos_4pos=0
    try:
        P_6pos_5pos=(N_6pos/495)/(N_5pos/496)
    except:
        P_6pos_5pos=0
    try:
        P_7pos_6pos=(N_7pos/494)/(N_6pos/495)
    except:
        P_7pos_6pos=0
    try:
        P_2neg_1pos=N_2neg_1pos/499
    except:
        P_2neg_1pos=0
    try:
        P_3neg_2pos=N_3neg_2pos/498
    except:
        P_3neg_2pos=0
    try:
        P_4neg_3pos=N_4neg_3pos/497
    except:
        P_4neg_3pos=0
    try:
        P_5neg_4pos=N_5neg_4pos/496
    except:
        P_5neg_4pos=0
    try:
        P_6neg_5pos=N_6neg_5pos/495
    except:
        P_6neg_5pos=0
    try:
        P_7neg_6pos=N_7neg_6pos/494
    except:
        P_7neg_6pos=0
    try:
        P_2pos_1neg=N_2pos_1neg/499
    except:
        P_2pos_1neg=0
    try:
        P_3pos_2neg=N_3pos_2neg/498
    except:
        P_3pos_2neg=0
    try:
        P_4pos_3neg=N_4pos_3neg/497
    except:
        P_4pos_3neg=0
    try:
        P_5pos_4neg=N_5pos_4neg/496
    except:
        P_5pos_4neg=0
    try:
        P_6pos_5neg=N_6pos_5neg/495
    except:
        P_6pos_5neg=0
    try:
        P_7pos_6neg=N_7pos_6neg/494
    except:
        P_7pos_6neg=0
#    print("P_2pos_1pos=%s %s %s %s %s %s\n"%(P_2pos_1pos,P_3pos_2pos,P_4pos_3pos,P_5pos_4pos,P_6pos_5pos,P_7pos_6pos))
#    print("P_2neg_1pos=%s %s %s %s %s %s\n"%(P_2neg_1pos,P_3neg_2pos,P_4neg_3pos,P_5neg_4pos,P_6neg_5pos,P_7neg_6pos))
#    print("P_2pos_1neg=%s %s %s %s %s %s\n"%(P_2pos_1neg,P_3pos_2neg,P_4pos_3neg,P_5pos_4neg,P_6pos_5neg,P_7pos_6neg))
#    
    
    avg2/=N_2pos
    avg3/=N_3pos
    avg4/=N_4pos
    try:
        avg5/=N_5pos
    except:
        avg5=0
    try:
        avg6/=N_6pos
    except:
        avg6=0
    
    try:
        avg7/=N_7pos
    except:
        avg7=0
    for j in range(2,30):
        if(j==2):#cols-353
            sheet2.cell(j,cols-353).value=P_2pos_1pos
        elif(j==3):
            sheet2.cell(j,cols-353).value=P_3pos_2pos
        elif(j==4):
            sheet2.cell(j,cols-353).value=P_4pos_3pos
        elif(j==5):
            sheet2.cell(j,cols-353).value=P_5pos_4pos
        elif(j==6):
            sheet2.cell(j,cols-353).value=P_6pos_5pos
        elif(j==7):
            sheet2.cell(j,cols-353).value=P_7pos_6pos
        elif(j==8):
            sheet2.cell(j,cols-353).value=P_2neg_1pos
        elif(j==9):
            sheet2.cell(j,cols-353).value=P_3neg_2pos
        elif(j==10):
            sheet2.cell(j,cols-353).value=P_4neg_3pos
        elif(j==11):
            sheet2.cell(j,cols-353).value=P_5neg_4pos
        elif(j==12):
            sheet2.cell(j,cols-353).value=P_6neg_5pos
        elif(j==13):
            sheet2.cell(j,cols-353).value=P_7neg_6pos
        elif(j==14):
            sheet2.cell(j,cols-353).value=P_2pos_1neg
        elif(j==15):
            sheet2.cell(j,cols-353).value=P_3pos_2neg
        elif(j==16):
            sheet2.cell(j,cols-353).value=P_4pos_3neg
        elif(j==17):
            sheet2.cell(j,cols-353).value=P_5pos_4neg
        elif(j==18):
            sheet2.cell(j,cols-353).value=P_6pos_5neg
        elif(j==19):
            sheet2.cell(j,cols-353).value=P_7pos_6neg
        elif(j==20):
            sheet2.cell(j,cols-353).value=P_pos_pos_neg
        elif(j==21):
            sheet2.cell(j,cols-353).value=P_neg_pos_neg
        elif(j==22):
            sheet2.cell(j,cols-353).value=P_pos_neg_pos
        elif(j==23):
            sheet2.cell(j,cols-353).value=P_neg_neg_pos
    sheet2.cell(24,cols-353).value=avg2
    sheet2.cell(25,cols-353).value=avg3
    sheet2.cell(26,cols-353).value=avg4
    sheet2.cell(27,cols-353).value=avg5
    sheet2.cell(28,cols-353).value=avg6
    sheet2.cell(29,cols-353).value=avg7
    
wb.save('C:\\Users\\Ajit\\excels\\'+today+'.xlsx')    
#returns of volume
today=datetime.datetime.today().strftime('%d-%m-%Y')
wb=openpyxl.load_workbook('C:\\Users\\Ajit\\excels\\'+today+'.xlsx',data_only=True)
sheets=wb.sheetnames
sheet1=wb[sheets[3]]
sheets=wb.sheetnames
sheet2=wb[sheets[4]]
for cols in range (2,353):
    for rows in range (2,524):
        try:
            returns=(float(sheet1.cell(rows,cols).value)-float(sheet1.cell(rows+1,cols).value))/float(sheet1.cell(rows+1,cols).value)
        except:
            returns=0
        sheet2.cell(rows,cols).value=returns


wb.save('C:\\Users\\Ajit\\excels\\'+today+'.xlsx') 
    
###VOLUME CONDITIONAL

today=datetime.datetime.today().strftime('%d-%m-%Y')

wb=openpyxl.load_workbook('C:\\Users\\Ajit\\excels\\'+today+'.xlsx',data_only=True)
sheets=wb.sheetnames
sheet1=wb[sheets[4]]#1 or 4
sheet2=wb[sheets[5]]#2 or 5
col_names=[]
for rows in sheet1.iter_rows(min_row=1,max_row=1,min_col=2,max_col=352):#355,706 or 2,352
    for cell in rows:
        col_names.append(cell.value)
j=1
for names in col_names:
    sheet2.cell(1,j).value=names
    j=j+1
sheet2.cell(2,1).value="P(2|1)"
sheet2.cell(3,1).value="P(3|2)"
sheet2.cell(4,1).value="P(4|3)"
sheet2.cell(5,1).value="P(5|4)"
sheet2.cell(6,1).value="P(6|5)"
sheet2.cell(7,1).value="P(7|6)"
sheet2.cell(8,1).value="P(2-|1)"
sheet2.cell(9,1).value="P(3-|2)"
sheet2.cell(10,1).value="P(4-|3)"
sheet2.cell(11,1).value="P(5-|4)"
sheet2.cell(12,1).value="P(6-|5)"
sheet2.cell(13,1).value="P(7-|6)"
sheet2.cell(14,1).value="P(2|1-)"
sheet2.cell(15,1).value="P(3|2-)"
sheet2.cell(16,1).value="P(4|3-)"
sheet2.cell(17,1).value="P(5|4-)"
sheet2.cell(18,1).value="P(6|5-)"
sheet2.cell(19,1).value="P(7|6-)"
sheet2.cell(20,1).value="P(++-)"
sheet2.cell(21,1).value="P(-+-)"
sheet2.cell(22,1).value="P(+-+)"
sheet2.cell(23,1).value="P(--+)"
sheet2.cell(24,1).value="avg2_1"
sheet2.cell(25,1).value="avg3_2"
sheet2.cell(26,1).value="avg4_3"
sheet2.cell(27,1).value="avg5_4"
sheet2.cell(28,1).value="avg6_5"
sheet2.cell(29,1).value="avg7_6"
sheet2.cell(30,1).value="P(2-|1-)"
sheet2.cell(31,1).value="P(3-|2-)"
sheet2.cell(32,1).value="P(4-|3-)"
sheet2.cell(33,1).value="P(5-|4-)"
sheet2.cell(34,1).value="P(6-|5-)"
sheet2.cell(35,1).value="P(7-|6-)"
for cols in range(2,353):#355,706
    returns=[]
    for rows in range(2,502):
        a=str(sheet1.cell(row=rows,column=cols).value)
        a=a.replace(',','')
        a=float(a)
        returns.append(a)
    returns.reverse()
    
    avg2=0
    avg3=0
    avg4=0
    avg5=0
    avg6=0
    avg7=0
    
    N_1pos=0
    
    for num in returns:
        if(num>0):
            N_1pos+=1
    
    
    
    
    N_2pos=0
    
    i=1#P(2|1)
    for num in returns:
        if(i<500):
            if(returns[i]>0 and num>0):
                N_2pos+=1
                avg2=avg2+returns[i]
        i=i+1
    
    i=2#P(3|2)
    N_3pos=0
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-1]>0 and returns[i]>0):
                N_3pos+=1
                avg3=avg3+returns[i]
        i=i+1
            
    N_4pos=0
    i=3
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]>0):
                N_4pos+=1
                avg4=avg4+returns[i]
        i=i+1
    N_5pos=0
    i=4
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]>0):
                N_5pos+=1
                avg5=avg5+returns[i]
        i=i+1
    N_6pos=0
    i=5
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-4]>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]>0):
                N_6pos+=1
                avg6+=returns[i]
        i=i+1
    N_7pos=0
    i=6
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-5]>0 and returns[i-4]>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]>0):
                N_7pos+=1
                avg7+=returns[i]
        i=i+1
    #NUM-1NEG,NUM-2NEG,NUM-3NEG,NUM-4NEG,NUM5NEG,NUM6NEG,NUM7NEG
    N_1neg=0
    for num in returns:
        if(num<=0):
            N_1neg+=1
    
    
    
    
    N_2neg=0
    
    i=1#P(2|1)
    for num in returns:
        if(i<500):
            if(returns[i]<=0 and num<=0):
                N_2neg+=1
                
        i=i+1
    
    i=2#P(3|2)
    N_3neg=0
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-1]<=0 and returns[i]<=0):
                N_3neg+=1
                
        i=i+1
            
    N_4neg=0
    i=3
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]<=0):
                N_4neg+=1
                
        i=i+1
    N_5neg=0
    i=4
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-3]<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]<=0):
                N_5neg+=1
        i=i+1
    N_6neg=0
    i=5
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-4]<=0 and returns[i-3]<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]<=0):
                N_6neg+=1
        i=i+1
    N_7neg=0
    i=6
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-5]<=0 and returns[i-4]<=0 and returns[i-3]<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]<=0):
                N_7neg+=1
        i=i+1
    
#    print("N_1pos=%s %s %s %s %s %s %s"%(N_1pos,N_2pos,N_3pos,N_4pos,N_5pos,N_6pos,N_7pos))
#    ###############################################################################################################
    
    
    N_2pos_1neg=0
    i=1
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i]>0):
                N_2pos_1neg+=1
        i=i+1
    
    N_3pos_2neg=0
    i=2
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-1]<=0 and returns[i]>0):
                N_3pos_2neg+=1
        i=i+1
    
    N_4pos_3neg=0
    i=3
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]>0):
                N_4pos_3neg+=1
        i=i+1
    
    N_5pos_4neg=0
    i=4
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-3]<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]>0):
                N_5pos_4neg+=1
        i=i+1
    N_6pos_5neg=0
    i=5
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-4]<=0 and returns[i-3]<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]>0):
                N_6pos_5neg+=1
        i=i+1
    N_7pos_6neg=0
    i=6
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-5]<=0 and returns[i-4]<=0 and returns[i-3]<=0 and returns[i-2]<=0 and returns[i-1]<=0 and returns[i]>0):
                N_7pos_6neg+=1
        i=i+1
#    print("N_2pos_1neg=%s %s %s %s %s %s"%(N_2pos_1neg,N_3pos_2neg,N_4pos_3neg,N_5pos_4neg,N_6pos_5neg,N_7pos_6neg))
#    ########################################################################################################
    
    N_2neg_1pos=0
    i=1
    for num in returns:
        if(i<500):
            if(num>0 and returns[i]<=0):
                N_2neg_1pos+=1
        i=i+1
    
    N_3neg_2pos=0
    i=2
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-1]>0 and returns[i]<=0):
                N_3neg_2pos+=1
        i=i+1
    
    N_4neg_3pos=0
    i=3
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]<=0):
                N_4neg_3pos+=1
        i=i+1
    
    N_5neg_4pos=0
    i=4
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]<=0):
                N_5neg_4pos+=1
        i=i+1
    N_6neg_5pos=0
    i=5
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-4]>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]<=0):
                N_6neg_5pos+=1
        i=i+1
    N_7neg_6pos=0
    i=6
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-5]>0 and returns[i-4]>0 and returns[i-3]>0 and returns[i-2]>0 and returns[i-1]>0 and returns[i]<=0):
                N_7neg_6pos+=1
        i=i+1
#    print("N_2neg_1pos=%s %s %s %s %s %s"%(N_2neg_1pos,N_3neg_2pos,N_4neg_3pos,N_5neg_4pos,N_6neg_5pos,N_7neg_6pos))
#   
    i=2#P(3|2)
    N_pos_pos_neg=0
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-1]>0 and returns[i]<=0):
                N_pos_pos_neg+=1
        i=i+1
    try:
        P_pos_pos_neg=N_pos_pos_neg/498
    except:
        P_pos_pos_neg=0

    i=2#P(3|2)
    N_neg_pos_neg=0
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-1]>0 and returns[i]<=0):
                N_neg_pos_neg+=1
        i=i+1
    try:
        P_neg_pos_neg=N_neg_pos_neg/498
    except:
        P_neg_pos_neg=0
    i=2#P(3|2)
    N_pos_neg_pos=0
    for num in returns:
        if(i<500):
            if(num>0 and returns[i-1]>0 and returns[i]>0):
                N_pos_neg_pos+=1
        i=i+1
    try:
        P_pos_neg_pos=N_pos_neg_pos/498
    except:
        P_pos_neg_pos=0
    i=2#P(3|2)
    N_neg_neg_pos=0
    for num in returns:
        if(i<500):
            if(num<=0 and returns[i-1]<=0 and returns[i]>0):
                N_neg_neg_pos+=1
        i=i+1
    try:
        P_neg_neg_pos=N_neg_neg_pos/498
    except:
        P_neg_neg_pos=0
#################################################################################################################
    try:
        P_2pos_1pos=(N_2pos/499)/(N_1pos/500)
    except:
        P_2pos_1pos=0
    try:
        P_3pos_2pos=(N_3pos/498)/(N_2pos/499)
    except:
        P_3pos_2pos=0
    try:
        P_4pos_3pos=(N_4pos/497)/(N_3pos/498)
    except:
        P_4pos_3pos=0
    try:
        
        P_5pos_4pos=(N_5pos/496)/(N_4pos/497)
    except:
        P_5pos_4pos=0
    try:
        P_6pos_5pos=(N_6pos/495)/(N_5pos/496)
    except:
        P_6pos_5pos=0
    try:
        P_7pos_6pos=(N_7pos/494)/(N_6pos/495)
    except:
        P_7pos_6pos=0
    try:
        P_2neg_1pos=N_2neg_1pos/499
    except:
        P_2neg_1pos=0
    try:
        P_3neg_2pos=N_3neg_2pos/498
    except:
        P_3neg_2pos=0
    try:
        P_4neg_3pos=N_4neg_3pos/497
    except:
        P_4neg_3pos=0
    try:
        P_5neg_4pos=N_5neg_4pos/496
    except:
        P_5neg_4pos=0
    try:
        P_6neg_5pos=N_6neg_5pos/495
    except:
        P_6neg_5pos=0
    try:
        P_7neg_6pos=N_7neg_6pos/494
    except:
        P_7neg_6pos=0
    try:
        P_2pos_1neg=N_2pos_1neg/499
    except:
        P_2pos_1neg=0
    try:
        P_3pos_2neg=N_3pos_2neg/498
    except:
        P_3pos_2neg=0
    try:
        P_4pos_3neg=N_4pos_3neg/497
    except:
        P_4pos_3neg=0
    try:
        P_5pos_4neg=N_5pos_4neg/496
    except:
        P_5pos_4neg=0
    try:
        P_6pos_5neg=N_6pos_5neg/495
    except:
        P_6pos_5neg=0
    try:
        P_7pos_6neg=N_7pos_6neg/494
    except:
        P_7pos_6neg=0
    try:
        P_2neg_1neg=(N_2neg/499)/(N_1neg/500)
    except:
        P_2neg_1neg=0
    try:
        P_3neg_2neg=(N_3neg/498)/(N_2neg/499)
    except:
        P_3neg_2neg=0
    try:
        P_4neg_3neg=(N_4neg/497)/(N_3neg/498)
    except:
        P_4neg_3neg=0
    try:
        P_5neg_4neg=(N_5neg/496)/(N_4neg/497)
    except:
        P_5neg_4neg=0
    try:
        P_6neg_5neg=(N_6neg/495)/(N_5neg/496)
    except:
        P_6neg_5neg=0
    try:
        P_7neg_6neg=(N_7neg/494)/(N_6neg/495)
    except:
        P_7neg_6neg=0
#    print("P_2pos_1pos=%s %s %s %s %s %s\n"%(P_2pos_1pos,P_3pos_2pos,P_4pos_3pos,P_5pos_4pos,P_6pos_5pos,P_7pos_6pos))
#    print("P_2neg_1pos=%s %s %s %s %s %s\n"%(P_2neg_1pos,P_3neg_2pos,P_4neg_3pos,P_5neg_4pos,P_6neg_5pos,P_7neg_6pos))
#    print("P_2pos_1neg=%s %s %s %s %s %s\n"%(P_2pos_1neg,P_3pos_2neg,P_4pos_3neg,P_5pos_4neg,P_6pos_5neg,P_7pos_6neg))
#    
    
    
    try:
        avg2/=N_2pos
    except:
        avg2=0
    try:
        avg3/=N_3pos
    except:
        avg3=0
    try:
        avg4/=N_4pos
    except:
        avg4=0
    try:
        avg5/=N_5pos
    except:
        avg5=0
    try:
        avg6/=N_6pos
    except:
        avg6=0
    
    try:
        avg7/=N_7pos
    except:
        avg7=0
    for j in range(2,30):
        if(j==2):#cols-353
            sheet2.cell(j,cols).value=P_2pos_1pos
        elif(j==3):
            sheet2.cell(j,cols).value=P_3pos_2pos
        elif(j==4):
            sheet2.cell(j,cols).value=P_4pos_3pos
        elif(j==5):
            sheet2.cell(j,cols).value=P_5pos_4pos
        elif(j==6):
            sheet2.cell(j,cols).value=P_6pos_5pos
        elif(j==7):
            sheet2.cell(j,cols).value=P_7pos_6pos
        elif(j==8):
            sheet2.cell(j,cols).value=P_2neg_1pos
        elif(j==9):
            sheet2.cell(j,cols).value=P_3neg_2pos
        elif(j==10):
            sheet2.cell(j,cols).value=P_4neg_3pos
        elif(j==11):
            sheet2.cell(j,cols).value=P_5neg_4pos
        elif(j==12):
            sheet2.cell(j,cols).value=P_6neg_5pos
        elif(j==13):
            sheet2.cell(j,cols).value=P_7neg_6pos
        elif(j==14):
            sheet2.cell(j,cols).value=P_2pos_1neg
        elif(j==15):
            sheet2.cell(j,cols).value=P_3pos_2neg
        elif(j==16):
            sheet2.cell(j,cols).value=P_4pos_3neg
        elif(j==17):
            sheet2.cell(j,cols).value=P_5pos_4neg
        elif(j==18):
            sheet2.cell(j,cols).value=P_6pos_5neg
        elif(j==19):
            sheet2.cell(j,cols).value=P_7pos_6neg
        elif(j==20):
            sheet2.cell(j,cols).value=P_pos_pos_neg
        elif(j==21):
            sheet2.cell(j,cols).value=P_neg_pos_neg
        elif(j==22):
            sheet2.cell(j,cols).value=P_pos_neg_pos
        elif(j==23):
            sheet2.cell(j,cols).value=P_neg_neg_pos
    sheet2.cell(24,cols).value=avg2
    sheet2.cell(25,cols).value=avg3
    sheet2.cell(26,cols).value=avg4
    sheet2.cell(27,cols).value=avg5
    sheet2.cell(28,cols).value=avg6
    sheet2.cell(29,cols).value=avg7
    sheet2.cell(30,cols).value=P_2neg_1neg
    sheet2.cell(31,cols).value=P_3neg_2neg
    sheet2.cell(32,cols).value=P_4neg_3neg
    sheet2.cell(33,cols).value=P_5neg_4neg
    sheet2.cell(34,cols).value=P_6neg_5neg
    sheet2.cell(35,cols).value=P_7neg_6neg
    
wb.save('C:\\Users\\Ajit\\excels\\'+today+'.xlsx')

wb=openpyxl.load_workbook('C:\\Users\\Ajit\\excels\\'+today+'.xlsx')
sheets=wb.sheetnames
ws=wb[sheets[1]]
ws_new=wb.create_sheet('ols_result')
i=1
row=6
for col in range(708,758):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=13
for col in range(708,724):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=20
for col in range(708,724):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=27
for col in range(708,720):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=34
for col in range(708,737):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1

row=41
for col in range(708,718):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=48
for col in range(708,723):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=55
for col in range(708,723):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=62
for col in range(708,738):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=69
for col in range(708,733):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=76
for col in range(708,718):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1

row=83
for col in range(708,723):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=90
for col in range(708,723):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=97
for col in range(708,738):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=104
for col in range(708,718):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=111
for col in range(708,720):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=118
for col in range(708,718):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
row=125
for col in range(708,738):
    if(float(ws.cell(row,col).value)>=1.96):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,1).value=ws.cell(row-1,col).value
        i=i+1
###################################################osl less than 1.96
i=1
row=6
for col in range(708,758):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=13
for col in range(708,724):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=20
for col in range(708,724):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=27
for col in range(708,720):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=34
for col in range(708,737):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1

row=41
for col in range(708,718):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=48
for col in range(708,723):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=55
for col in range(708,723):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=62
for col in range(708,738):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=69
for col in range(708,733):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=76
for col in range(708,718):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1

row=83
for col in range(708,723):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=90
for col in range(708,723):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=97
for col in range(708,738):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=104
for col in range(708,718):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=111
for col in range(708,720):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=118
for col in range(708,718):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
row=125
for col in range(708,738):
    if(float(ws.cell(row,col).value)<(-1.96)):
        print(ws.cell(row-1,col).value)
        ws_new.cell(i,2).value=ws.cell(row-1,col).value
        i=i+1
wb.save('C:\\Users\\Ajit\\excels\\'+today+'.xlsx')


